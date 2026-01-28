"""
Write the standard annual report for the climberdb database. The script uses 
an Excel template to place query results
"""
from copy import (
	copy,
	deepcopy
)
from openpyxl import (
	load_workbook,
	Workbook
)
from openpyxl.styles import (
	Alignment,
	Border,
	Side
)
from openpyxl.styles.styleable import StyleArray
from openpyxl.utils.cell import get_column_letter
from pandas import (
	concat,
	DataFrame,
	isna,
	Series
)
from re import (
	match, 
	sub
)
from typing import Optional
from typing_extensions import Self

from climberdb_utils import query_db

side_style = Side(color='000000', style='thin')
_default_border_style = Border(
	top=side_style, 
	left=side_style, 
	right=side_style, 
	bottom=side_style
)

class ExcelReportRenderer:

	_ANCHOR_REGEX = r'\{\{.+?\}\}'
	_DEFAULT_ROW_HEIGHT = 26
	_DEFAULT_ROW_WIDTH = 90 / 7 #openpyxl col dimensions are 7 pixels/unit 🤷

	def __init__(self, template_path, tag_sep='::', attribute_sep=';'):
		self.template = load_workbook(template_path)
		self.src_ws = self.template.active
		self.tag_sep = tag_sep
		self.attribute_sep = attribute_sep

		self.cells, self.row_heights = self._snapshot_cells()
		self.col_widths = self._snapshot_column_widths()
		self.anchors = self._load_anchors()
		self.merges = self._snapshot_merges()

	# ───────── Snapshot phase ─────────

	def _snapshot_cells(self):
		rows = []
		heights = {}
		widths = {}

		for r in range(1, self.src_ws.max_row + 1):
			heights[r] = self.src_ws.row_dimensions[r].height
			for c in range(1, self.src_ws.max_column + 1):
				cell = self.src_ws.cell(r, c)
				value = (
					sub(self._ANCHOR_REGEX, cell.value, '') 
					if match(self._ANCHOR_REGEX, cell.value or '') 
					else cell.value
				)
				rows.append({
					"row": r,
					"col": c,
					"value": value,
					"original_cell": cell 
				})

		return DataFrame(rows), heights


	def _snapshot_column_widths(self):
		ws = self.src_ws

		widths = {}
		for col in range(1, ws.max_column + 1):
		    letter = get_column_letter(col)
		    dim = ws.column_dimensions.get(letter)
		    widths[letter] = dim.width if dim and dim.width else self._DEFAULT_ROW_WIDTH
		
		return widths

	def _snapshot_merges(self):
		return DataFrame(
			[m.bounds for m in self.src_ws.merged_cells],
			columns=["start_column","start_row","end_column","end_row"]
		)

	def _load_anchors(self):
		anchors = []
		for r in range(1, self.src_ws.max_row + 1):
			for c in range(1, self.src_ws.max_column + 1):
				v = self.src_ws.cell(r, c).value or ''
				if match(self._ANCHOR_REGEX, v):
					parts = sub('[{}]', '', v).split(self.tag_sep)
					tag = parts[0]
					attrs = parts[1].split(self.attribute_sep) if len(parts) > 1 else []
					anchors.append({
						"tag": tag, "row": r, "col": c,
						"include_header": "include_header" in attrs,
						"include_index": "exclude_index" not in attrs
					})
		return DataFrame(anchors)

	# ───────── Compile phase ─────────

	def _set_cell_style(self, dst_cell, src_cell):
		
		if not isna(src_cell) and src_cell.has_style:
			dst_cell.font = copy(src_cell.font)
			dst_cell.border = copy(src_cell.border)
			dst_cell.fill = copy(src_cell.fill)
			dst_cell.number_format = copy(src_cell.number_format)
			dst_cell.protection = copy(src_cell.protection)
			alignment = src_cell.alignment

			# cell.alignment is not directly copy-able
			dst_cell.alignment = Alignment(
			    horizontal = alignment.horizontal,
			    vertical   = alignment.vertical,
			    wrap_text  = alignment.wrap_text
			)


	def render_report(self, tables, values, output_path):

		wb = Workbook()
		ws = wb.active

		logical_cells = self.cells.copy()
		logical_merges = self.merges.copy()
		original_anchors = self.anchors.copy()

		def shift_rows(start, delta):
			logical_cells.loc[logical_cells.row >= start, 'row'] += delta
			logical_merges.loc[logical_merges.start_row >= start,
				['start_row','end_row']] += delta
			self.anchors.loc[self.anchors.row >= start, 'row'] += delta
			self.cells.loc[self.cells.row >= start, 'row'] += delta

		# process anchors in template order
		for anchor_index in self.anchors.sort_values('row').index:
			a = self.anchors.loc[anchor_index].squeeze()
			new_cells = []
			if a.tag in tables:
				df = tables[a.tag]

				needed = len(df) - 1#
				original_row = original_anchors.loc[anchor_index, 'row']
				if needed > 0:
					shift_rows(a.row + 1, needed)

				r = a.row
				if a.include_header:
					# Remove the anchor text
					ws.cell(r, 1).value = ''
					
					for i, h in enumerate(df.columns):
						col = a.col + i + 1
						new_cells.append({
							"row": r, 
							"col": col, 
							"value": h, 
							"original_cell": self.src_ws.cell(original_row, col)
							})
					r += 1
					original_row += 1

				for rowdata in df.itertuples(index=a.include_index):
					#for i, v in enumerate(rowdata):
					for col in range(1, self.src_ws.max_column + 1): 
						value = rowdata[col - a.col] if col >= a.col and col <= len(rowdata) else None
						new_cells.append({
							"row": r, 
							"col": col, 
							"value": value, 
							"original_cell": self.src_ws.cell(original_row, col)
							})
					r += 1
				logical_cells = concat(
					[logical_cells, DataFrame(new_cells)],
					ignore_index=True
				)
			elif a.tag in values:
				indexer = (
					(logical_cells.row == a.row) & (logical_cells.col == a.col)
				)
				logical_cells.loc[indexer, 'value'] = values[a.tag]
				logical_cells.loc[indexer, 'original_cell'] = self.src_ws.cell(original_anchors.loc[anchor_index, 'row'], a.col)

		# ───────── Materialize sheet ─────────

		row_heights_set = []
		for _, cell_data in logical_cells.iterrows():
			dst_cell = ws.cell(cell_data.row, cell_data.col, cell_data.value)
			src_cell = cell_data.original_cell
			self._set_cell_style(dst_cell, src_cell)

			# Set row heigt if it hasn't yet been set
			src_row = src_cell.row
			dst_row = cell_data.row
			if not dst_row in row_heights_set:
				ws.row_dimensions[dst_row].height = self.row_heights[src_row] or self._DEFAULT_ROW_HEIGHT

		# # Set index to be able to find cells by row/col combo. It's safe to
		# # 	alter the dataframe here because it's not used after this
		# logical_cells = logical_cells.set_index(['row', 'col'])
		# # Find all the row, col indices of merged cells to exclude them 
		# #	from the list of blanks
		# merged_cell_indices = [
		# 	[*c] # decompose tuple into list
		# 	for mc in self.src_ws.merged_cells 
		# 	for c in mc.cells
		# ]
		# blank_cells = logical_cells.loc[
		# 	isna(logical_cells.value) &
		# 	~logical_cells.index.isin(merged_cell_indices)
		# ]

		for col, width in self.col_widths.items():
			ws.column_dimensions[col].width = width

		#import pdb; pdb.set_trace()
		for _, m in logical_merges.iterrows():
			ws.merge_cells(**m)

		wb.save(output_path)




class AnnualSummary:

	_DENALI_CODE = 1
	_FORAKER_CODE = 2
	
	_GENDER_CODES = {
		1: 'female',
		2: 'male'
	}
	
	_TEMPLATE_SQL = '''
		SELECT 
			{select}
		FROM (
			SELECT DISTINCT ON ({distinct_on})
			*
			FROM {schema}.registered_climbs_view
			WHERE year = {year} 
			{where}
		) _
		GROUP BY {group_by}
	'''
	
	_TOTAL_CLIMBERS_DESCRIPTION = 'Total Climbers (does not include canceled)'

	def __init__(self, year, schema='public'):
		if not isinstance(year, int):
			raise ValueError('year must be int')

		self.year = year
		self.schema = schema

		self.mountain_code_str = f"{self._DENALI_CODE}, {self._FORAKER_CODE}"
		self.default_params = {
			'year': year, 
			'schema': schema, 
			'select': 'mountain_code, count(expedition_member_id) AS value',
			'distinct_on': 'expedition_member_id, mountain_code',
			'where': f' AND mountain_code in ({self.mountain_code_str})',
			'group_by': 'mountain_code'
		}
		# Country/state queries rely on total from snapshot. Set it to default 
		#	to an empty dataframe so _build_states_countries_of_origin() can at 
		#	least check that it's valid
		self.snapshot = DataFrame();

		# Define a dict to hold mappings of Excel template anchors (all defined 
		#	in cells in the template as by "{{<tag_name>}}" ) to map them to 
		#	their values
		self.tags = {}


	def _load_reference_tables(self):
		self.mountain_codes = (
			DataFrame(query_db({'tables': ['mountain_codes']}))
				.set_index('code')
		)
		self.route_codes = (
			DataFrame(query_db({'tables': ['route_codes']}))
		)
		self.mountain_names = self.mountain_codes\
			.loc[
				self.mountain_codes.index.isin([
					self._DENALI_CODE, 
					self._FORAKER_CODE
				]), 
				'name'
			]
		self.states = (
			DataFrame(query_db({'tables': ['state_codes']}))
				.set_index('code')
		)
		self.countries = ( 
			DataFrame(query_db({'tables': ['country_codes']}))
				.set_index('code')
		)
		self.us_country_code = (
			self.countries
				.loc[self.countries.short_name == 'US']
				.index[0]
		)
		self.denali_name = self.mountain_names[self._DENALI_CODE]


	def _query_snapshot(
			self, 
			where:str='', 
			totals_for_percent:Series=Series(dtype=int)
		) -> DataFrame:
		"""
		Run a query for a single row of the 'Snapshot of the Season' section
		"""
		params = {**self.default_params}
		params['where'] += where

		data = DataFrame(
			query_db({
				'sql': self._TEMPLATE_SQL.format(**params)
			}),
			columns=['mountain_code', 'value']
		)

		# Set the index to just 0s so all rows collapse to one with .pivot
		data.index = [0] * len(data)
		# Transform data
		data = (data
			# drop any columns that aren't Denali or Foraker
			.loc[data.mountain_code.isin(self.mountain_names.index)]
			# replace codes with names
			.replace({'mountain_code': self.mountain_names})
			# make mountain names the columns (produces 1 row)
			.pivot(columns='mountain_code', values='value')
			# if any mountains were missing, add them as columns
			.reindex(columns=self.mountain_names)
			# fill any missing values with 0
			.fillna(0)
		)

		# Calculate scalar total of both mountains
		total = data.sum(axis=1).squeeze()
		data['Total'] = total

		# Percentages for total climbers per mountain is calculated as the 
		#   proportion of climbers on each mountain relative to the total 
		#   number of climbers. All other percentages are calculated as the 
		#   proportion of each climber type relative to the number of climbers 
		#   on each mountain
		if len(totals_for_percent) == 0:
			totals_for_percent = data.copy()
			totals_for_percent[data.columns] = total

		if len(data):
			for mountain_name_ in self.mountain_names:
				percent =  (
					  data[mountain_name_] 
					/ totals_for_percent[mountain_name_] 
					* 100
				).round(1).fillna(0).astype(str) + '%'
				data[mountain_name_ + ' % of Total'] = percent

		return data


	def _build_snapshot(self):

		# Get total climbers per mountain
		total_climbers = self._query_snapshot()
		totals_for_percent = (
			total_climbers
				.loc[:, self.mountain_names]
				.squeeze(axis=1)
		)

		# Climbers on guided trips (including guides)
		guided_climbers = self._query_snapshot(
			where=' AND guide_company_code <> -1 AND special_group_type_code <> 3',
			totals_for_percent=totals_for_percent
		)
		# Indpendent climbers
		independent_climbers = self._query_snapshot(
			where=' AND guide_company_code = -1 AND special_group_type_code <> 3',
			totals_for_percent=totals_for_percent
		)
		# NPS rangers and volunteers
		nps_climbers = self._query_snapshot(
			where=' AND special_group_type_code = 3',
			totals_for_percent=totals_for_percent
		)
		# Summits
		summited_climbers = self._query_snapshot(
			where=" AND route_was_summited",
			totals_for_percent=totals_for_percent
		)
		# Female climbers
		female_climbers = self._query_snapshot(
			where=' AND sex_code = 1',
			totals_for_percent=totals_for_percent
		)
		# Female summits
		female_summits = self._query_snapshot(
			where=" AND sex_code = 1 AND summited = 'Yes'",
			totals_for_percent=summited_climbers
				.loc[:, self.mountain_names]
				.squeeze(axis=1)
				.fillna(0)
		)
		
		snapshot = concat([
			total_climbers,
			guided_climbers,
			independent_climbers,
			nps_climbers,
			female_climbers,
			summited_climbers,
			female_summits
		])
		snapshot['description'] = [
			self._TOTAL_CLIMBERS_DESCRIPTION,
			'Guided Climbers including Guides',
			'Independent Climbers',
			'NPS Mountaineering Rangers/Volunteers on Mountain',
			'Women Climbers (included in above numbers)',
			'Total Summits',
			'Women Summits (included in total summits)'
		]

		# Query guided vs independent Denali summit percentage
		params = self.default_params.copy()
		params['where'] = '''
			AND special_group_type_code <> 3 
			AND coalesce(guide_company_code, -1) <> -1
			AND mountain_code = 1
			AND summited = 'Yes'
		'''
		guided_summits = query_db(
			{'sql': self._TEMPLATE_SQL.format(**params)}
		)

		params['where'] = params['where'].replace('<> -1', '= -1')
		independent_summits = query_db(
			{'sql': self._TEMPLATE_SQL.format(**params)}
		)

		if len(guided_summits):
			guided_summits = guided_summits[0]
			guided_summit_percent = round(
				  guided_summits['value'] 
				/ guided_climbers.loc[0, self.denali_name] 
				* 100
			)
			snapshot = concat([
				snapshot,
				DataFrame([{
					self.denali_name: guided_summit_percent,
					'description': f'{self.denali_name} Guided Summit Rate'
				}])
			])
		if len(independent_summits):
			independent_summits = independent_summits[0]
			independent_summit_percent = round(
				  independent_summits['value'] 
				/ independent_climbers.loc[0, self.denali_name] 
				* 100
			)
			snapshot = concat([
				snapshot,
				DataFrame([{
					self.denali_name: independent_summit_percent,
					'description': f'{self.denali_name} Non-Guided Summit Rate'
				}])
			])

		# Get first and last summit dates. 
		# TODO: consider just getting summit count by day to get first and last and save as instance attribute. Then get the top 3 summit dates from that dataframe
		sql = f''' 
			SELECT 
				to_char(min(summit_date), 'FMMonth DD') AS Earliest,
				to_char(max(summit_date), 'FMMonth DD') AS Latest,
				mountain_code
			FROM 
				{self.schema}.registered_climbs_view
			WHERE year = {self.year}
				AND mountain_code IN ({self.mountain_code_str})
			GROUP BY mountain_code
		'''
		summit_dates = (DataFrame(query_db({'sql': sql}))
			.set_index('mountain_code')
			.T
		)

		summit_dates.rename(columns=self.mountain_names, inplace=True)
		summit_dates['description'] = (
			summit_dates
				.index
				.astype(str)
				.str
				.capitalize() 
			+ ' Summit Date'
		)
		if len(summit_dates):
			snapshot = concat([
				snapshot,
				summit_dates.reset_index(drop=True)
			]).reindex(columns=[
				'description', 
				self.denali_name, 
				'Foraker', 
				'Total', 
				self.denali_name + ' % of Total', 
				'Foraker % of Total'
			]).set_index('description')

		self.snapshot = snapshot


	def _get_top_states_countries(self, data, n=4):
		"""
		Helper method to sort queries of states and contries of origin.
		Don't just get the 4 most popular states. Get the top 4 values 
		and then get any states with a number of climbers that matches any 
		of those 4 values. This is so that if there are, say, 2 climbers 
		from 6 states, but 2 is the 4th highest value, all 6 states are 
		included
		"""

		top = {}
		# Loop through mountains in order of their code, so Denali will be first
		for mountain_name_ in self.mountain_names.sort_index():
			top_counts = (data
				.dropna(subset=mountain_name_)
				.sort_values(mountain_name_, ascending=False)
				.head(n)[mountain_name_]
			)
			top[mountain_name_] = (
				data[mountain_name_]
					.loc[data[mountain_name_].isin(top_counts)]
			)

		return DataFrame(top).sort_values(self.denali_name, ascending=False)


	def _build_states_countries_of_origin(self):
		"""
		Build the tables summarizing where climbers are from
		"""

		# Make sure total climbers has been queried first (in _build_snapshot())
		if (
			not len(self.snapshot) and 
			not len(self.snapshot.loc[
					self.snapshot.index == self._TOTAL_CLIMBERS_DESCRIPTION
				])
			):
			raise RuntimeError(
				'_build_snapshot() method must be called before'
				' _build_states_countries_of_origin()'
			)
		total_climbers = self.snapshot.loc[self._TOTAL_CLIMBERS_DESCRIPTION]

		# Set default params for queries of where climbers are from
		default_params = self.default_params.copy()
		default_params['where'] = (
				f' AND mountain_code IN ({self.mountain_code_str}) ' 
				f' AND country_code = {self.us_country_code} '
		)
		default_params['distinct_on'] = 'climber_id, mountain_code'
		default_params['select'] = 'count(*), state_code, mountain_code'
		default_params['group_by'] = 'state_code, mountain_code'
		
		us_total_params = default_params.copy()
		us_total_params['select'] = 'count(*), 1 AS a'
		us_total_params['distinct_on'] = 'climber_id'
		us_total_params['group_by'] = 'a'
		us_total_sql = self._TEMPLATE_SQL.format(**us_total_params)
		us_total = query_db({'sql': us_total_sql})[0]['count']
		us_percent = round(
			us_total / total_climbers.Total.squeeze() * 100,
			1
		)
		self.us_total = us_total
		self.us_percent = str(us_percent) + '%'

		us_climbers_params = default_params.copy()
		us_climbers_params['where'] += ' AND state_code IS NOT NULL'
		us_climbers_sql = self._TEMPLATE_SQL.format(**us_climbers_params)
		us_climbers = (
			DataFrame(
				query_db({'sql': us_climbers_sql})
			).pivot(columns='mountain_code', index='state_code', values='count')
			.rename(columns=self.mountain_names)
		)

		self.n_states = len(us_climbers)

		top_states = self._get_top_states_countries(us_climbers)
		self.top_states = top_states.rename(index=self.states.name)

		# Non-US climbers
		self.intl_total = total_climbers.Total.squeeze() - us_total
		self.intl_percent = str(round(100 - us_percent, 1)) + '%'
		
		intl_params = default_params.copy()
		intl_params['where'] = (
			intl_params['where']
				.replace('country_code =', 'country_code <>')
				+ ' AND country_code IS NOT NULL'
		)
		intl_params['select'] = 'count(*), country_code, mountain_code'
		intl_params['group_by'] = 'country_code, mountain_code'
		intl_sql = self._TEMPLATE_SQL.format(**intl_params)
		intl_climbers = (
			DataFrame(
				query_db({'sql': intl_sql})
			).pivot(columns='mountain_code', index='country_code', values='count')
			.rename(columns=self.mountain_names)
		)
		self.n_countries = len(intl_climbers)
		top_countries = self._get_top_states_countries(intl_climbers, n=6)
		# Get the code values of the top Denali countries
		top_denali_countries = (
			top_countries
				.dropna(subset=self.denali_name)
				.index
		)
		self.top_countries = top_countries.rename(index=self.countries.name)

		# Get all countries that had 3 or less climbers
		less_than_3 = intl_climbers.loc[intl_climbers[self.denali_name] <= 3]
		self.less_than_3_total = str(
				int(less_than_3[self.denali_name].sum())
			) + ' climbers'
		less_than_3_countries = less_than_3.index

		# Get all other counteis
		other_countries = intl_climbers.loc[
			~intl_climbers.index.isin(
				top_denali_countries.tolist() + less_than_3.index.tolist()
			)
		]
		self.less_than_3_list = ', '.join(
			less_than_3
				.rename(index=self.countries.name)
				.index
					.sort_values()
		)
		self.other_countries_list = ', '.join(
			other_countries
				.rename(index=self.countries.name)
				.index
					.sort_values()
		)
		self.other_counters_total = str(
				int(other_countries[self.denali_name].sum())
			) + ' climbers'


	def _build_demographics(self):
		"""
		Query climber age and gender
		"""
		default_params = self.default_params.copy()
		default_params['select'] = '''
			avg(age)::int AS average, 
			min(age) AS youngest, 
			max(age) AS oldest, 
			mountain_code, 
			sex_code
		'''
		default_params['distinct_on'] = 'climber_id'
		default_params['group_by'] = 'mountain_code, sex_code'

		def split_by_gender(data):
			to_combine = []
			for code, gender in self._GENDER_CODES.items():
				split = (
					data.loc[data.sex_code == code]
						.set_index('mountain_code')
						.drop(columns='sex_code')
						.rename(columns=self.mountain_names)
						.T
				)
				split.index += f'_{gender}_age'
				to_combine.append(split)
			
			return concat(to_combine)

		age_sql = self._TEMPLATE_SQL.format(**default_params)
		climber_age = DataFrame(query_db({'sql': age_sql}))
		age_by_gender = split_by_gender(climber_age)
		
		summit_age_params = default_params.copy()
		summit_age_params['where'] += ' AND route_was_summited'
		summit_age_sql = self._TEMPLATE_SQL.format(**summit_age_params)
		summit_age = DataFrame(query_db({'sql': summit_age_sql}))
		summit_age_by_gender = split_by_gender(summit_age)
		summit_age_by_gender.index += '_summited'

		demographics_order = {
			'average_female_age': 'Average Age of Female Climbers',
			'average_male_age': 'Average Age of Male Climbers',
			'oldest_female_age': 'Oldest Female Climber',
			'oldest_female_age_summited': 'Oldest Female to Summit',
			'oldest_male_age': 'Oldest Male Climber',
			'oldest_male_age_summited': 'Oldest Male to Summit',
			'youngest_female_age': 'Youngest Female Climber',
			'youngest_female_age_summited': 'Youngest Female to Summit',
			'youngest_male_age': 'Youngest Male Climber',
			'youngest_male_age_summited': 'Youngest Male to Summit'
		}
		self.demographics = (
			concat([
				age_by_gender, 
				summit_age_by_gender
			]).loc[demographics_order.keys()]
			.rename(
				columns=self.mountain_names,
				index=demographics_order
			)
		)


		def process_guides_rangers(sql, data_type='Guide'):
			data = (
				DataFrame(query_db({'sql': sql}))
					.set_index('sex_code')
					.rename(index=self._GENDER_CODES)
			)
			data['average'] = (
				data['average']
					.astype(float) # type might be object
					.round()
					.astype(int)
			)
			total = data['count'].sum()
			data['percent'] = (
				data['count'] / total * 100
			).round().astype(int).astype(str) + '%'

			total_index_str = f'Total Number of {data_type}s'
			data.loc[total_index_str] = total

			guide_genders = (
				data.loc[:, ['percent']]
				.rename(
					index={
						'female': f'Percent Female {data_type}s', 
						'male': f'Percent Male {data_type}s'
					},
					columns={'percent': 0}
				).loc[[
					total_index_str,
					f'Percent Female {data_type}s',
					f'Percent Male {data_type}s',
				]]
			)
			
			guide_ages = data.loc[:, ['average']].rename(
				index={
					'female': f'Average Age of Female {data_type}',
					'male': f'Average Age of Male {data_type}',
				},
				columns={'average': 0}
			)
			return concat([guide_genders, guide_ages])\
				.drop_duplicates()

		# Guides
		guide_params = default_params.copy()
		guide_params['select'] = '''
			round(avg(age)) AS average, 
			count(*), 
			sex_code
		'''
		guide_params['group_by'] = 'sex_code'
		guide_params['where'] += ' AND is_guiding'
		guide_sql = self._TEMPLATE_SQL.format(**guide_params)
		self.guides = process_guides_rangers(guide_sql, 'Guide')

		# Rangers
		ranger_params = guide_params.copy()
		ranger_params['where'] = ranger_params['where'].replace(
			' AND is_guiding', 
			' AND special_group_type_code = 3'
		)
		ranger_sql = self._TEMPLATE_SQL.format(**ranger_params)
		self.rangers = process_guides_rangers(ranger_sql, 'Ranger')


	def _build_summit_dates(self):
		# Summits per month and top 3 summit dates
		summit_by_month_params = self.default_params.copy()
		summit_by_month_params['select'] = 'month, count(*)'
		summit_by_month_params['group_by'] = 'month'
		summit_by_month_params['distinct_on'] = 'expedition_member_id'
		summit_by_month_params['where'] += ' AND summit_date IS NOT NULL AND mountain_code = 1'
		summit_by_month_sql = self._TEMPLATE_SQL.format(**summit_by_month_params)
		self.summits_by_month = (DataFrame(
				query_db({'sql': summit_by_month_sql})
			).set_index('month')
			.reindex(index=['May', 'June', 'July'], fill_value=0)
		)

		top_summit_dates_params = summit_by_month_params.copy()
		top_summit_dates_params['select'] = '''
			to_char(summit_date, 'FMMonth DD') as summit_date, 
			count(*)
		'''
		top_summit_dates_params['group_by'] = 'summit_date'
		top_summit_dates_params['distinct_on'] = 'expedition_member_id'
		top_summit_dates_sql = self._TEMPLATE_SQL\
			.format(**top_summit_dates_params)
		self.top_summit_dates = (
			DataFrame(query_db({'sql': top_summit_dates_sql}))
				.sort_values('count', ascending=False)
				.head(3)
				.set_index('summit_date')
		)


	def _build_routes(self) -> None:
		"""
		Query count of attempts and summits for each route climbed
		"""
		routes_params = self.default_params.copy()
		routes_params['distinct_on'] = 'route_name, expedition_member_id'
		routes_params['select'] = 'mountain_code, route_name, summited, count(*)'
		routes_params['group_by'] = 'mountain_code, route_name, summited'
		routes_sql = self._TEMPLATE_SQL.format(**routes_params)
		routes = DataFrame(query_db({'sql': routes_sql}))

		def get_route_counts(mountain_code):
			mountain_routes = (
				routes
					.loc[routes.mountain_code == mountain_code]
					.drop(columns='mountain_code')
					.pivot(
						columns='summited', 
						index='route_name', 
						values='count'
					).rename(columns={'Yes': 'Summits'})
					.reindex(columns=['Attempts', 'Summits', 'No'])
					.fillna(0)
			)
			mountain_routes['Attempts'] = mountain_routes.sum(axis=1)
			mountain_routes = mountain_routes.drop(columns='No')
			mountain_routes['Success Rate'] = (
				mountain_routes.Summits / mountain_routes.Attempts * 100
			).round().astype(int)

			return mountain_routes.sort_values('Attempts', ascending=False)
		
		self.denali_routes = get_route_counts(self._DENALI_CODE)
		self.foraker_routes = get_route_counts(self._FORAKER_CODE)


	def _build_trip_length(self) -> None:
		"""
		Query average overall, guided, and NPS trip lengths
		"""
		trip_length_params = self.default_params.copy()
		avg_select = 'round(avg(trip_length_days)) AS avg_trip_length'
		trip_length_params['select'] = f'''
			{avg_select},
			mountain_code,
			CASE 
				WHEN guide_company_code = -1 
				THEN 'Non-Guided' 
				ELSE 'Guided' 
			END AS guided 
		'''
		trip_length_params['distinct_on'] = 'expedition_id, mountain_code'
		trip_length_params['group_by'] = 'mountain_code, guided'
		trip_length_sql = self._TEMPLATE_SQL.format(**trip_length_params)
		trip_length_all = (
			DataFrame(query_db({'sql': trip_length_sql}))
				.pivot(
					columns='guided', 
					index='mountain_code', 
					values='avg_trip_length'
				).rename(index='All ' + self.mountain_names + ' trips')
		)
		trip_summit_params = trip_length_params.copy()
		trip_summit_params['where'] += ' AND route_was_summited'
		trip_length_summit_sql = self._TEMPLATE_SQL.format(**trip_summit_params)
		trip_length_summit = (
			DataFrame(query_db({'sql': trip_length_summit_sql}))
				.pivot(
					columns='guided', 
					index='mountain_code', 
					values='avg_trip_length'
				).rename(
					index=self.mountain_names + ' trips with successful summits'
				)
		)
		trip_length_nps_params = trip_length_params.copy()
		trip_length_nps_params['select'] = f'''
			{avg_select}, 
			mountain_code
		'''
		trip_length_nps_params['group_by'] = 'mountain_code'
		trip_length_nps_params['where'] += ' AND special_group_type_code = 3'
		trip_length_nps_sql = self._TEMPLATE_SQL.format(**trip_length_nps_params)
		trip_length_nps = (
			DataFrame(query_db({'sql': trip_length_nps_sql}))
				.set_index('mountain_code')
				.rename(
					columns={'avg_trip_length': 'Non-Guided'},
					index=self.mountain_names + ' NPS patrols'
				)
		)

		self.trip_length = concat([
			trip_length_all,
			trip_length_summit,
			trip_length_nps
		]).reindex(columns=['Non-Guided', 'Guided'])


	def run_report(self) -> Self:
		"""
		Run all the queries
		"""	

		self._load_reference_tables()
		self._build_snapshot()
		self._build_states_countries_of_origin()
		self._build_demographics()
		self._build_summit_dates()
		self._build_routes()
		self._build_trip_length()

		self.tags = {
			'tables': {
				'snapshot_table':       self.snapshot,
				'top_states_table':     self.top_states,
				'top_countries_table':  self.top_countries,
				'demographics_table':   self.demographics,
				'guides_table':         self.guides,
				'rangers_table':        self.rangers,
				'monthly_summits_table': self.summits_by_month,
				'top_summit_dates_table': self.top_summit_dates,
				'denali_routes_table':  self.denali_routes,
				'foraker_routes_table': self.foraker_routes,
				'trip_length_table':    self.trip_length
			},
			'values': {
				'n_states':             self.n_states,
				'us_total':				self.us_total,
				'us_percent':			self.us_percent,
				'n_countries':          self.n_countries,
				'international_total':	self.intl_total,
				'international_percent': self.intl_percent,
				'countries_w_less_than_3': self.less_than_3_list,
				'less_than_3_total':    self.less_than_3_total,
				'other_countries':      self.other_countries_list,
				'other_countries_total': self.other_counters_total,
				'denali_name': 			self.denali_name
			}
		}

		# Return self so we can do AnnualSummary().run_report() in one line
		return self