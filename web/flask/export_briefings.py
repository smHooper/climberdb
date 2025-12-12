"""
Create an Excel file of briefing appointment schedules for each day in a given
range. Breifing appointment details come from the climberdb frontend.
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.comments import Comment
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
from typing import Dict


PRINT_WIDTH_INCHES = 8.5 #for 8.5x11
PRINT_HEIGHT_INCHES = 11
# The actual default print area height appears to be 8" and 1032 px. So there's 
#	apparently 108 vertical pixels per inch. However, the units in Excel rows 
#	are 3/4 the number of pixels. Horizontal ratios are slightly different 🤷
WIDTH_FUDGE_FACTOR = 0.14058 
HEIGHT_FUDGE_FACTOR = 3/4
DEFAULT_LEFT_MARGIN = 0.75
DEFAULT_TOP_MARGIN = 1.5 # 1" margin + 0.5" header/footer
X_PIXELS_PER_INCH = 89.1428
Y_PIXELS_PER_INCH = 108
PRINT_WIDTH = (
	X_PIXELS_PER_INCH 
		* (PRINT_WIDTH_INCHES - (DEFAULT_LEFT_MARGIN * 2)) 
		* WIDTH_FUDGE_FACTOR
)
PRINT_HEIGHT = (
	Y_PIXELS_PER_INCH 
	* (
		PRINT_HEIGHT_INCHES 
		- ((DEFAULT_TOP_MARGIN - .1) * 2)
	) 
	* HEIGHT_FUDGE_FACTOR
)


def add_briefing(
		sheet: Worksheet, 
		briefing_info: dict, 
		day_start_column:int=1
	) -> None:

	"""
	Add a worksheet to the given workbook that shows the briefing schedule for a single day

	Parameters
	____________
	sheet: An openpyxl.Worksheet to write to
	briefing_info: dict with information for a single briefing including the 
		row/column indices it should occupy, the expedition name, additional text, and colors
	day_start_column: an integer offset to add to the column indices. Since 
		multiple days might be shown on the same worksheet and column indices
		are relative to the day, the offset allows calculation of the absolute 
		column incidces

	Returns
	___________
	None
	"""
	start_row, start_column, end_row, end_column = briefing_info['cell_indices']
	
	# Add 1 to account for header row
	start_row += 1
	end_row += 1

	is_single_row = end_row == start_row 
	
	# Briefing info start and end columns are relative to that day only, so add 
	#	the number of columns from previous days
	start_column += day_start_column
	end_column += day_start_column

	# Merge the top row
	# Occasionally, the briefings overlap even though they shouldn't. If so,
	#	just skip this one
	try:
		title_cell = sheet.cell(
			row=start_row, 
			column=start_column, 
			value=briefing_info['expedition_name']
		)
	except Exception as e:
		print(briefing_info['cell_indices'])
		return
	
	# Create styles
	fill_color = briefing_info['background_color']
	fill = PatternFill(
		fill_type='solid', 
		start_color=fill_color, 
		end_color=fill_color
	)
	side_style = Side(
		border_style='thick', 
		color='B86000'
	)
	box_border = Border(
		top=side_style, 
		left=side_style, 
		right=side_style, 
		bottom=side_style
	)
	top_border = Border(
		top=side_style, 
		left=side_style, 
		right=side_style
	)
	bottom_border = Border(
		left=side_style, 
		right=side_style, 
		bottom=side_style
	)
	header_alignment = Alignment(
		horizontal='left', 
		vertical='center', 
		indent=2.0, 
		wrapText=True
	)
	body_alignment = Alignment(
		horizontal='left', 
		vertical='top', 
		indent=2.0, 
		wrapText=True
	)

	# Setting border style on merged cells doesn't work (it only sets the
	# border on the first cell), so set 
	#	each cell's style separately, then merge
	for i in range(start_column, end_column + 1):
		cell = sheet.cell(row=start_row, column=i)
		cell.border = box_border if is_single_row else top_border
		cell.alignment = header_alignment
		cell.fill = fill
		cell.font = Font(bold=True, color=briefing_info['text_color'])
	sheet.merge_cells(
		start_row=start_row, 
		start_column=start_column, 
		end_row=start_row, 
		end_column=end_column
	)

	comment = Comment(str(briefing_info.get('comment')), 'Briefing Export')
	if briefing_info['comment']:
		cell.comment = comment

	# Merge the rest if the briefing is more than one row
	if not is_single_row:
		info_top_row_index = start_row + 1
		upper_left_cell = sheet.cell(
			row=info_top_row_index, 
			column=start_column, 
			value=briefing_info['briefing_text']
		)
		
		# Setting border style on the merged cells doesn't work, so set the
		#	style on each cell, then merge
		row_iterator = sheet.iter_rows(
			min_row=info_top_row_index, 
			min_col=start_column,
			max_row=end_row, 
			max_col=end_column
		)
		for row in row_iterator:
			for cell in row:
				cell.border = bottom_border
				cell.fill = fill
				cell.alignment = body_alignment
				cell.font = Font(color=briefing_info['text_color'])

		try:
			sheet.merge_cells(
				start_row=info_top_row_index, 
				start_column=start_column, 
				end_row=end_row, 
				end_column=end_column
			)
		except:
			pass

		if briefing_info['comment']:
			upper_left_cell.comment = comment


def addDateHeader(
		sheet: Worksheet, 
		dt: datetime, 
		start_column:int, 
		end_column:int
	) -> None:
	"""
	Helper function to add the date header for each day in a worksheet

	Parameters
	____________
	sheet: An openpyxl.Worksheet to write to
	dt: datetime object to convert the date into a formatted string
	start_column: the absolute column index for the start of this day
	end_column: the absolute column index for the end of this day

	Returns
	___________
	None
	"""
	cell = sheet.cell(
		row=1, 
		column=start_column,
		value = dt.strftime('%A, %B %#d')
	)
	cell.alignment = Alignment(horizontal='center', vertical='center')
	sheet.merge_cells(
		start_row=1, 
		start_column=start_column, 
		end_row=1, 
		end_column=end_column
	)


def create_sheet(
		workbook: Workbook, 
		briefings: Dict, 
		time_slots: list[str]
	) -> None:
	"""
	Add a worksheet to the given workbook that shows the briefing schedule for a single day or multiple days

	Parameters
	____________
	workbook: An openpyxl.Workbook to write to
	date_str: ISO date like 2023-05-10 to use as the worksheet name
	briefings: a nested dictionary where keys are ISO date strings and values 	
		are a list of dictionaries. Each dict in the list contains information 
		for a single briefing including the row/column indices it should 
		occupy, the expedition name, additional text, and colors
	time_slots: the labels for time slots in the schedule

	Returns
	___________
	None
	
	"""
	dates = list(briefings.keys())
	sheet_name = dates[0] # for simplicity, just name it after the first one 
	sheet = workbook.create_sheet(sheet_name)
	_time_label_width = 7

	# Set the "background" cell borders to show a faint horizontal line at the 
	#	top of each hour. Do this before adding briefing appointments to the 
	#	schedule so the borders of the appointments will overwrite the 
	#	"background" lines
	bottom_side_border = Side(border_style='thin', color='D9D9D9')
	right_side_border = Side(border_style='thin', color='000000')
	bottom_only_style = Border(
		bottom=bottom_side_border
	)
	right_only_style = Border(
		right=right_side_border
	)
	bottom_and_right_style = Border(
		bottom=bottom_side_border,
		right=right_side_border
	)
	
	row_height = PRINT_HEIGHT / (len(time_slots) * 2)

	# Get the column bounds for the end (right side) of each day. This is based 
	#	on the number of concurrent briefings, each of which needs to fit in a 
	#	column of the Excel sheet
	current_column = 1 # first column index is 1, not 0
	column_ends = [1]
	column_widths = []
	for _, briefing_list in briefings.items():
		if len(briefing_list) == 0:
			n_columns = 1
		else:
			n_columns = max([
				briefing_info['cell_indices'][-1] 
				for briefing_info 
				in briefing_list
			])
		current_column += n_columns
		column_ends.append(current_column)
		column_widths.append(n_columns)

	# Add time slots
	# cell indices are 1-based and should start on the second row
	time_slot_rows = [i * 2 + 2 for i in range(len(time_slots))]

	for i, row_index in enumerate(range(1, time_slot_rows[-1] + 1)):
		# Set row height
		sheet.row_dimensions[row_index].height = row_height

		# Time labels should be on the hour but each row represents a half hour
		cell = sheet.cell(row=row_index, column=1)

		if row_index in time_slot_rows:
			label = time_slots[i // 2] # floor divide will return int, not float
			cell.value = label
			cell.alignment = Alignment(vertical='top')
			cell.border = bottom_and_right_style
		else:
			cell.border = right_only_style

		# Set "background" border style
		for column_index in range(2, column_ends[-1] + 1):
			cell = sheet.cell(row=row_index, column=column_index)
			needs_bottom_border = row_index in time_slot_rows
			needs_right_border = column_index in column_ends
			if needs_bottom_border and needs_right_border:
				border_style = bottom_and_right_style
			elif needs_bottom_border:
				border_style = bottom_only_style
			elif needs_right_border:
				border_style = right_only_style
			else:
				# it's neither a time slot row or the right-most column in a day
				continue
			cell.border = border_style

	# Set time label column width
	sheet.column_dimensions['A'].width = _time_label_width

	# Add briefings and set columns for each date
	day_start_column = column_ends[0] + 1
	for day_index, (date_str, briefing_list) in enumerate(briefings.items()):
		
		day_column_width = column_widths[day_index]
		day_end_column = day_start_column + day_column_width - 1
		try:
			dt = datetime.strptime(date_str, '%Y-%m-%d')
		except:
			dt = None

		# If there are no briefings scheduled, leave it blank and exit
		if len(briefing_list) == 0:
			if dt != None:
				addDateHeader(sheet, dt, day_start_column, day_end_column)

		for briefing_info in briefing_list:
			add_briefing(
				sheet, 
				briefing_info, 
				day_start_column=day_start_column - 1
			)

		# Set col width
		# Subtract for time label column
		schedule_width_px = PRINT_WIDTH - _time_label_width 
		# Then floor divide by the number columns this day uses
		column_width = schedule_width_px // day_column_width
		for c in range(day_start_column, day_end_column + 1): # +1 because range is exclusive
			sheet.column_dimensions[get_column_letter(c)].width = column_width

		# Set the column heading for this day as the formated date		
		if dt != None:
			addDateHeader(sheet, dt, day_start_column, day_end_column)

		day_start_column += day_column_width

	# Freeze top row and left column
	sheet.freeze_panes = sheet.cell(row=2, column=2)
	sheet.print_title_cols = 'A:A'


def briefings_to_excel(data: Dict, output_dir: str) -> str:
	"""
	Create an Excel doc of briefing appointments per day from a dictionary of 
	appointment details. Each day will be a separate sheet in the doc.

	Parameters
	____________
	data: Dictionary with keys 'time_slots' and 'briefings'. time_slots is a 
		list of times to label the schedule. Briefings is dictionary of with a key/value pair for each day. The value of each day is a list of 
		dictionaries, and each dictionary represents a different appointment. The data object is therefore in the format:
			{
				time_slots: ['7am', '8am', ...],
				sheets: [
					{
						2023-05-10: [{<breifing1 details}, {<breifing2 details}, ...],
						2023-05-11: [{<breifing1 details}, {<breifing2 details}, ...],
						...
					},
					...
				]
			}
	output_dir: Path to write the Excel doc to. The filename will be generated
		from the min and max dates specified in data['sheets']
	
	Returns
	__________
	str: filename of the Excel doc

	"""
	workbook = Workbook()
	del workbook['Sheet']

	if not 'time_slots' in data:
		raise KeyError('"time_slots" not given in data')

	if not 'sheets' in data:
		raise KeyError('"briefings" not given in data')

	if not os.path.isdir(output_dir):
		raise IOError(f'Cannot write output to output_dir "{output_dir}" because it does not exist')

	time_slots = data['time_slots']
	sheets = data['sheets']

	# Create a sheet per day
	min_date = 'a'
	max_date = ''
	for briefings in sheets:
		create_sheet(workbook, briefings, time_slots)
		briefing_dates = [date_str for date_str in briefings.keys()]
		min_date = min(*briefing_dates, min_date)
		max_date = max(*briefing_dates, max_date)

	# Write the Excel file
	briefing_dates = briefings.keys()
	filename = f'''briefing_schedule_{min_date}_to_{max_date}.xlsx'''
	excel_path = os.path.join(output_dir, filename)
	workbook.save(excel_path)

	return filename